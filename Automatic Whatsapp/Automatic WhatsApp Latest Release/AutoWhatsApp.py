import tkinter as tk 
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from fontTools.ttLib import TTFont
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import time, os, datetime, webbrowser, pandas, pyautogui, schedule



#Special Font
SF = TTFont('font/SFCompact.ttf')


#DataBase Options

##Openpyxl
#Create A Workbook
wb = load_workbook('data/data.xlsx')

#Set Active Worksheet
ws = wb.active

#Grab A Column Of Data
phonecolumn = ws["B"]
messagecolumn = ws["C"]
namecolumn = ws["A"]


##Pandas
df = pandas.read_excel("data/data.xlsx")
data_dict = df.to_dict('list')

phone = data_dict['Phone']
name = data_dict['Name']
message = data_dict['Message']

Comp = zip(phone,message)

#Main Window Options
root = Tk()
root.title('Auto WhatsApp')
root.geometry("550x400")
root.resizable (False, False)
root.configure(bg="#1c1c1e")
root.iconbitmap('icon/AWhatsApp.ico')

#Creating Interface
interface = ttk.Notebook(root)
interface.pack()

#Interface Options and Creating Tabs
main_frame = Frame(interface, width=550, height=450, bg="#1c1c1e")
second_frame = Frame(interface, width=550, height=450, bg="#1c1c1e")

def hide():
    interface.forget(1)

def refresh():
    label1.config(text="You Have No Message To Send.")
    functionlabel2.config(text='')


main_frame.pack(fill="both", expand=1,anchor='center')
second_frame.pack( fill="both",expand=1,anchor='center')

#Tab Options
interface.add(main_frame, text="Home")
interface.add(second_frame, text="New")

#Main Frame Items

##Clock
timelabel = Label(main_frame,bg='#1c1c1e',fg='#7D7AFF')

def clock():
    dt = datetime.datetime.now().strftime("%H:%M:%S")
    timelabel.config(text=dt)
    timelabel.after(1000, clock)

timelabel.pack(pady=55)
timelabel.config(font=("SF", 25))
clock()

##Labels
global label1
label1 = Label(main_frame,text="You Have No Message To Send.",bg='#1c1c1e',fg='white')
label1.config(font=("SF"))
label1.pack()

global functionlabel2
functionlabel2 = Label(main_frame, text='*Please click help before starting.' + "\n" + '*You need to login WhatsApp Web to synchronize.',bg='#1c1c1e',fg='#67b3e6')
functionlabel2.pack(pady=50)
functionlabel2.config(font=("SF", 10))

##This Is For Stabilizing The Page Width 
line = Label(main_frame,text="_____________________________________________________________",bg='#1c1c1e',fg='white',font='SF')
line.pack(pady='1000')



#Creating Menu
wmenu = Menu(main_frame)
root.config(menu=wmenu)

#Menu Commands
##Navigating To Our Second Frame
def gomessage():
    interface.add(second_frame, text="New")
    interface.select(1) 

##Navigating To Our Main Frame
def gomain():
    interface.select(0)

##I Let You To Send 
def youcansend():
    sendok()

##Navigating To Our Github Repo
def gowebsite():
    webbrowser.open_new_tab("https://github.com/abdullahcat/Automatic-Messages-With-WhatsApp")
def gousermanual():
    webbrowser.open_new_tab("https://github.com/abdullahcat/Automatic-Messages-With-WhatsApp/blob/main/UserManuel.md")
    
#Menu Items

##File Menu
file_menu = Menu(wmenu,tearoff=0)
wmenu.add_cascade(label='File', menu=file_menu)
file_menu.add_command(label='Exit',command=root.quit)
file_menu.add_command(label='Refresh',command=refresh)

##Edit Menu
edit_menu = Menu(wmenu,tearoff=0)
wmenu.add_cascade(label='Edit', menu=edit_menu)
edit_menu.add_command(label='New Message',command=gomessage)
edit_menu.add_command(label='Send The Message',command=youcansend)


##Help Menu
help_menu = Menu(wmenu,tearoff=0)
wmenu.add_cascade(label='Help', menu=help_menu)
help_menu.add_command(label='User Manual',command=gousermanual)
help_menu.add_separator()
help_menu.add_command(label='About Auto WhatsApp',command=gowebsite)
help_menu.add_command(label='Github Repo',command=gowebsite)




#OUR COMMANDS
##Message Send Command Using PyAutoGUI
def sendbulk():

    for phone,message in Comp:
       time.sleep(20)
       webbrowser.open("https://web.whatsapp.com/send?phone="+phone+"&text="+message)

    if True:
        time.sleep(5)
        first=False

    width,height = pyautogui.size()
    pyautogui.click(width/2,height/2)

    time.sleep(7)
    pyautogui.press('enter')

    time.sleep(7)
    pyautogui.hotkey('ctrl', 'w')   


#Second Frame Tabs Items

##Second Frame Label
who = Label(second_frame,text="Who do you want to send message?", font='SF',fg='white',bg='#1c1c1e')
who.pack(pady=10)

    #ListBox Scrollbar
    #scrollbar_frame = Frame(second_frame)
    #list_scrollbar =Scrollbar(scrollbar_frame, orient=VERTICAL,bg="black")

##Listbox
contacts = Listbox(second_frame,height=5,bg="black",fg="white",font="SF",selectmode=MULTIPLE)
contacts.pack()

    ##Configure Scrollbar
    #list_scrollbar.config(command= contacts.yview)
    #list_scrollbar.pack(side=RIGHT, fill=Y)
    #scrollbar_frame.pack()
    
    #Items In Listbox
    #contacts.insert(END, "Tevfik")
    #contacts.insert(1, "Ece")
    #contacts.insert(0, "Mete")


##Transfering Excel Data to Listbox 
for item in namecolumn:
    contacts.insert(END, item.value)


##Second Frame Label
when = Label(second_frame,text="When do you want to send message?", font='SF',fg='white',bg='#1c1c1e')
when.pack(pady=8)

##Setting Time With Spinboxes
ghours_value = tk.StringVar(value=0)
ghour = ttk.Spinbox(second_frame,from_=0,to=23,wrap=True,width=5,state="readonly",textvariable=ghours_value)
ghour.pack(pady=10)

gminutes_value = tk.StringVar(value=0)
gminute = ttk.Spinbox(second_frame,from_=0,to=59,wrap=True,width=5,textvariable=gminutes_value)
gminute.pack()





##Functions For ListBox

def select():
    result = ''
    for item in contacts.curselection():
        result =  result + str(contacts.get(item)) + "\n " 

    label1.config(text="Waiting for the send message to " + result)
    funchour = ghours_value.get()
    funcminute = gminutes_value.get()
    functionlabel2.config(text= "Your message will send at " + str(funchour) + ":"  + str(funcminute))
    gomain()
    hide()
    

def sendok():
    while(1==1):
           if(ghours_value.get() == datetime.datetime.now().hour and
           gminutes_value.get() == datetime.datetime.now().minute) :
             sendbulk()
           break


##Buttons For Second Frame
CancelButton = Button(second_frame, text="Cancel", font='SF',bg='#ff2d55',fg='white',width=8,command=gomain)
CancelButton.pack(pady=10)

SaveButton = Button(second_frame, text="Save",font='SF',bg='#32ade6',fg='white',width=8,command=select)
SaveButton.pack()

root.mainloop()


